import openpyxl
from openpyxl.styles import PatternFill

class ExcelUtills():

    @staticmethod
    def getRowCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_row)

    @staticmethod
    def getColumnCount(file,sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return (sheet.max_column)

    @staticmethod
    def readData(file,sheetName,row,column):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row,column).value

    @staticmethod
    def writeData(file, sheetName, row, column, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row, column).value = data
        workbook.save(file)

    @staticmethod
    def fillGreenColour(file, sheetName, row, column):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        greenfill = PatternFill(start_color='60b212',
                                end_color='60b212',
                                fill_type='solid')
        sheet.cell(row, column).fill = greenfill
        workbook.save(file)

    @staticmethod
    def fillRedColour(file, sheetName, row, column):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetName]
        redfill = PatternFill(start_color='ff0000',
                                end_color='ff0000',
                                fill_type='solid')
        sheet.cell(row, column).fill = redfill
        workbook.save(file)